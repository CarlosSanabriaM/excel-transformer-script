import openpyxl
from openpyxl.styles import Font


def transform_excel(input_file, output_file, start_column_name):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Find the start column index
    start_column_index = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == start_column_name:
            start_column_index = col
            break

    if start_column_index is None:
        raise ValueError(f"Column '{start_column_name}' not found in the Excel file.")

    # Define the red font
    red_font = Font(color="FF0000")

    # Iterate over the rows starting from the second row
    for row in range(2, sheet.max_row + 1):
        col = start_column_index
        while col < sheet.max_column:
            int_value_cell = sheet.cell(row=row, column=col)
            perc_value_cell = sheet.cell(row=row, column=col + 1)

            # Check if the integer value is less than 100
            if int_value_cell.value is not None and int_value_cell.value < 100:
                int_value_cell.value = 0
                int_value_cell.font = red_font  # Set the text color to red
                perc_value_cell.value = 0
                perc_value_cell.font = red_font  # Set the text color to red

            col += 2  # Move to the next pair of columns

    # Save the modified workbook preserving the formatting
    wb.save(output_file)
    print(f"Modified Excel successfully saved as '{output_file}'")


if __name__ == "__main__":
    # Update these values
    input_excel_file = 'input.xlsx'  # Replace with your input file name
    output_excel_file = 'output.xlsx'  # Replace with your desired output file name
    start_column_name = 'First column to modify'  # Replace with your specific start column name

    transform_excel(input_excel_file, output_excel_file, start_column_name)
